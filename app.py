import streamlit as st
import zipfile, json, os, io, re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Brandlight Payroll → חשבשבת", page_icon="📊", layout="wide")

DEFAULT_ACCOUNTS = {"Gross Wages":"800000","Expense Reimbursement":"800001",
    "Employer Federal & State Taxes":"800002","Workers Compensation":"800003",
    "Employee Benefits":"800004","Administrative Fee":"800005",
    "401(k) ER Contribution":"800006","401(k) Establishment Fee":"800007"}
DEFAULT_CREDIT = "540001"
RANGE_NAME = "SALARIES"
INVOICE_FIELDS_ORDER = ["Gross Wages","Expense Reimbursement","Employer Federal & State Taxes",
    "Workers' Compensation","Employee Benefits","Administrative Fee",
    "Other: 401(k) ER Contribution","Other: 401(k) Establishment Fee"]
FIELD_TO_ACCOUNT = {"Gross Wages":"Gross Wages","Expense Reimbursement":"Expense Reimbursement",
    "Employer Federal & State Taxes":"Employer Federal & State Taxes",
    "Workers' Compensation":"Workers Compensation","Employee Benefits":"Employee Benefits",
    "Administrative Fee":"Administrative Fee","Other: 401(k) ER Contribution":"401(k) ER Contribution",
    "Other: 401(k) Establishment Fee":"401(k) Establishment Fee"}
HISTORY_FILE = "payroll_history.json"

def load_history():
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE,'r') as f: return json.load(f)
    return []
def save_history(h):
    with open(HISTORY_FILE,'w') as f: json.dump(h,f,indent=2,ensure_ascii=False)
def add_to_history(plist, history):
    existing = {h['invoice_number'] for h in history}
    added = 0
    for p in plist:
        if p['invoice_number'] not in existing:
            history.append({
                'pay_date':p.get('pay_date',''),'pay_date_hebrew':p.get('pay_date_hebrew',''),
                'invoice_number':p.get('invoice_number',''),
                'invoice_total':p['invoice'].get('total',0),
                'components':{k:v for k,v in (p['invoice'].get('components',{}) or {}).items()},
                'employees':p.get('employees',[]),
            })
            added += 1
    history.sort(key=lambda x: x.get('pay_date',''))
    return added

def read_pages(file_bytes):
    t = {}
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes),'r') as z:
            for tf in sorted([f for f in z.namelist() if f.endswith('.txt')],key=lambda x:int(x.replace('.txt',''))):
                t[int(tf.replace('.txt',''))] = z.read(tf).decode('utf-8',errors='replace')
            if t: return t
    except: pass
    try:
        from pypdf import PdfReader
        for i,page in enumerate(PdfReader(io.BytesIO(file_bytes)).pages):
            txt = page.extract_text() or ""
            if txt.strip(): t[i+1] = txt
    except ImportError:
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                for i,page in enumerate(pdf.pages):
                    txt = page.extract_text() or ""
                    if txt.strip(): t[i+1] = txt
        except: raise RuntimeError("Install pypdf: pip install pypdf")
    return t

def parse_invoice(text):
    fnames = ["Gross Wages","Expense Reimbursement","Employer Federal & State Taxes",
        "Workers' Compensation","Employee Benefits","Administrative Fee",
        "Other: 401(k) ER Contribution","Other: 401(k) Establishment Fee"]
    lines = [l.strip() for l in text.replace('\r','').split('\n') if l.strip()]
    result = {f:None for f in fnames}; total = None
    ff = []; al = []
    for i,line in enumerate(lines):
        for field in fnames:
            if line == field or line.startswith(field): ff.append((i,field))
        if re.match(r'^[\d,]+\.\d{2}$',line): al.append((i,float(line.replace(',',''))))
    if ff and al:
        last = ff[-1][0]; rel = [(i,a) for i,a in al if i > last]
        for i,(_,field) in enumerate(ff):
            if i < len(rel): result[field] = rel[i][1]
    if all(v is None for v in result.values()):
        for line in lines:
            for field in fnames:
                if field in line and result[field] is None:
                    a = re.findall(r'[\d,]+\.\d{2}',line)
                    if a: result[field] = float(a[-1].replace(',',''))
    if all(v is None for v in result.values()):
        full = ' '.join(lines)
        for field in fnames:
            m = re.search(re.escape(field)+r'\s+([\d,]+\.\d{2})',full)
            if m: result[field] = float(m.group(1).replace(',',''))
    for line in lines:
        if 'TOTAL INVOICE' in line:
            a = re.findall(r'[\d,]+\.\d{2}',line)
            if a: total = float(a[-1].replace(',',''))
        elif 'SUB-TOTAL' in line and total is None:
            a = re.findall(r'[\d,]+\.\d{2}',line)
            if a: total = float(a[-1].replace(',',''))
    if total is None:
        m = re.search(r'TOTAL\s+INVOICE\s+([\d,]+\.\d{2})',' '.join(lines))
        if m: total = float(m.group(1).replace(',',''))
    return {'components':result,'total':total}

def parse_employees(all_text):
    employees = []; cur_name = None; cur_id = None
    for pn in sorted(all_text.keys()):
        text = all_text[pn]
        if 'Payroll Register Detail' not in text: continue
        for line in text.replace('\r','').split('\n'):
            em = re.search(r'^(.+?)\s+Emp ID\s*:\s*(\w+)',line)
            if em:
                cur_name = em.group(1).strip(); cur_id = em.group(2).strip(); continue
            tp = re.search(r'Total Pay\s+[\d.]+\s+([\d,]+\.\d{2})\s+Total Taxes\s+([\d,]+\.\d{2})\s+Total Deductions\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})',line)
            if tp and cur_name:
                employees.append({'name':cur_name,'id':cur_id,
                    'gross_pay':float(tp.group(1).replace(',','')),
                    'taxes':float(tp.group(2).replace(',','')),
                    'ee_deductions':float(tp.group(3).replace(',','')),
                    'er_contributions':float(tp.group(4).replace(',',''))})
                cur_name = None; cur_id = None
    # Get full employer cost from Allocation Report
    ec = {}; ce = None
    for pn in sorted(all_text.keys()):
        text = all_text[pn]
        if 'Payroll Allocation Report' not in text or 'GRAND TOTALS' in text: continue
        for line in text.replace('\r','').split('\n'):
            for emp in employees:
                last = emp['name'].split(',')[0].strip().upper() if ',' in emp['name'] else emp['name'].split()[-1].upper()
                if len(last) > 2 and last in line.upper() and 'Total' not in line and 'GRAND' not in line: ce = emp['id']
            if ce and 'Contribution Total:' in line:
                m = re.search(r'Contribution Total:\s+([\d,]+\.\d{2})',line)
                if m and ce not in ec: ec[ce] = float(m.group(1).replace(',',''))
                ce = None
    for emp in employees:
        emp['employer_cost'] = ec.get(emp['id'], emp['er_contributions'])
        emp['total_cost'] = emp['gross_pay'] + emp['employer_cost']
    return employees

def extract_pdf_data(uploaded_file):
    data = {"employees":[],"invoice":{},"pay_date":"","invoice_number":""}
    fb = uploaded_file.read(); uploaded_file.seek(0)
    at = read_pages(fb)
    if not at: raise ValueError("Could not extract text")
    for pn in sorted(at.keys()):
        if 'TOTAL INVOICE' in at[pn] or 'SUB-TOTAL' in at[pn]:
            data['invoice'] = parse_invoice(at[pn]); break
    for pn in sorted(at.keys()):
        m = re.search(r'Pay Date\s+(\d{2}/\d{2}/\d{4})',at[pn])
        if m: data['pay_date'] = m.group(1); break
    for pn in sorted(at.keys()):
        m = re.search(r'Invoice(?:\s+#?\s*|\s+No\s+)(\d{7})',at[pn])
        if m: data['invoice_number'] = m.group(1); break
    if not data['invoice_number']:
        for pn in sorted(at.keys()):
            m = re.search(r'Invoice\s+(\d{7})',at[pn])
            if m: data['invoice_number'] = m.group(1); break
    data['employees'] = parse_employees(at)
    return data

def convert_date(us): p=us.split('/'); return f"{p[1]}/{p[0]}/{p[2]}" if len(p)==3 else us

def make_styles():
    hf=Font(name='Arial',bold=True,size=11,color='FFFFFF')
    hfill=PatternFill('solid',fgColor='2F5496')
    nf=Font(name='Arial',size=10); bf=Font(name='Arial',bold=True,size=10)
    tfill=PatternFill('solid',fgColor='E2EFDA')
    mfmt='#,##0.00'; pfmt='0.0%'
    tb=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    def sc(cell,font=None,fill=None,fmt=None,align=None):
        if font: cell.font=font
        if fill: cell.fill=fill
        if fmt: cell.number_format=fmt
        if align: cell.alignment=align
        cell.border=tb
    return hf,hfill,nf,bf,tfill,mfmt,pfmt,tb,sc

def generate_journal(plist, accounts, credit):
    wb=Workbook(); hf,hfill,nf,bf,tfill,mfmt,pfmt,tb,sc=make_styles()
    ws=wb.active; ws.title='פקודת יומן'
    ws.merge_cells('A1:H1')
    dates=[p['pay_date_hebrew'] for p in plist]
    ws['A1']=f'BRANDLIGHT INC. - פקודת יומן שכר | טווח: {RANGE_NAME} | {dates[0]} - {dates[-1]}' if dates else ''
    ws['A1'].font=Font(name='Arial',bold=True,size=13,color='2F5496'); ws['A1'].alignment=Alignment(horizontal='center')
    for c,h in enumerate(['תאריך','חשבון חובה 1','חשבון זכות 1','חשבון זכות 2','פרטים','אסמכתא','חובה מט"ח','זכות מט"ח'],1):
        sc(ws.cell(row=3,column=c,value=h),font=hf,fill=hfill,align=Alignment(horizontal='center',wrap_text=True))
    row=4
    for p in plist:
        comps=p['invoice'].get('components',{})
        for field in INVOICE_FIELDS_ORDER:
            amt=comps.get(field,0) or 0
            if amt==0: continue
            an=FIELD_TO_ACCOUNT.get(field,field); num=accounts.get(an,"")
            sc(ws.cell(row=row,column=1,value=p['pay_date_hebrew']),font=nf,fmt='@'); ws.cell(row=row,column=1).number_format='@'
            sc(ws.cell(row=row,column=2,value=num),font=nf,fmt='@'); ws.cell(row=row,column=2).number_format='@'; ws.cell(row=row,column=2).alignment=Alignment(horizontal='center')
            sc(ws.cell(row=row,column=3,value=credit),font=nf,fmt='@'); ws.cell(row=row,column=3).number_format='@'; ws.cell(row=row,column=3).alignment=Alignment(horizontal='center')
            sc(ws.cell(row=row,column=4,value=''),font=nf,fmt='@')
            sc(ws.cell(row=row,column=5,value=an),font=nf)
            sc(ws.cell(row=row,column=6,value=p['invoice_number']),font=nf,fmt='@'); ws.cell(row=row,column=6).number_format='@'; ws.cell(row=row,column=6).alignment=Alignment(horizontal='center')
            sc(ws.cell(row=row,column=7,value=amt),font=nf,fmt=mfmt)
            sc(ws.cell(row=row,column=8,value=amt),font=nf,fmt=mfmt)
            row+=1
    for c,w in enumerate([14,16,16,16,34,14,18,18],1): ws.column_dimensions[get_column_letter(c)].width=w
    return wb

def generate_summary(history):
    wb=Workbook(); hf,hfill,nf,bf,tfill,mfmt,pfmt,tb,sc=make_styles(); yr=datetime.now().year
    sh=['Pay Date','Invoice #','Gross Wages','Expense Reimb.','ER Fed&State Tax','Workers Comp','Emp Benefits','Admin Fee','401k ER Contrib','401k Est Fee','Total Invoice','Verify']

    # Sheet 1: סיכום לפי תקופה
    ws=wb.active; ws.title='סיכום לפי תקופה'; ws.merge_cells('A1:L1')
    ws['A1']=f'BRANDLIGHT INC. - Invoice Summary YTD {yr}'; ws['A1'].font=Font(name='Arial',bold=True,size=14,color='2F5496'); ws['A1'].alignment=Alignment(horizontal='center')
    for c,h in enumerate(sh,1): sc(ws.cell(row=3,column=c,value=h),font=hf,fill=hfill,align=Alignment(horizontal='center',wrap_text=True))
    for i,h in enumerate(history):
        r=4+i; comps=h.get('components',{})
        sc(ws.cell(row=r,column=1,value=h.get('pay_date_hebrew','')),font=nf)
        sc(ws.cell(row=r,column=2,value=h.get('invoice_number','')),font=nf)
        for ci,field in enumerate(INVOICE_FIELDS_ORDER): sc(ws.cell(row=r,column=3+ci,value=comps.get(field,0) or 0),font=nf,fmt=mfmt)
        sc(ws.cell(row=r,column=11,value=h.get('invoice_total',0)),font=bf,fmt=mfmt)
        sc(ws.cell(row=r,column=12,value=f'=IF(ABS(SUM(C{r}:J{r})-K{r})<0.02,"✓","✗")'),font=nf)
    tr=4+len(history)
    sc(ws.cell(row=tr,column=1,value='YTD TOTAL'),font=bf,fill=tfill); sc(ws.cell(row=tr,column=2),fill=tfill)
    for c in range(3,13): sc(ws.cell(row=tr,column=c,value=f'=SUM({get_column_letter(c)}4:{get_column_letter(c)}{tr-1})') if c<12 else sc(ws.cell(row=tr,column=c),fill=tfill),font=bf,fill=tfill,fmt=mfmt)
    for c in range(1,13): ws.column_dimensions[get_column_letter(c)].width=16

    # Sheet 2: עלות לפי עובד
    ws2=wb.create_sheet('עלות לפי עובד'); ws2.merge_cells('A1:H1')
    ws2['A1']=f'BRANDLIGHT INC. - Employee Cost by Period YTD {yr}'; ws2['A1'].font=Font(name='Arial',bold=True,size=14,color='2F5496'); ws2['A1'].alignment=Alignment(horizontal='center')
    for c,h in enumerate(['Pay Date','Invoice #','Employee ID','Employee Name','Gross Pay ($)','Employer Cost ($)','Total Cost ($)','Verify'],1):
        sc(ws2.cell(row=3,column=c,value=h),font=hf,fill=hfill,align=Alignment(horizontal='center',wrap_text=True))
    r=4
    for h in history:
        emps=h.get('employees',[])
        sr=r
        for emp in emps:
            sc(ws2.cell(row=r,column=1,value=h.get('pay_date_hebrew','')),font=nf)
            sc(ws2.cell(row=r,column=2,value=h.get('invoice_number','')),font=nf)
            sc(ws2.cell(row=r,column=3,value=emp.get('id','')),font=nf)
            sc(ws2.cell(row=r,column=4,value=emp.get('name','')),font=nf)
            sc(ws2.cell(row=r,column=5,value=emp.get('gross_pay',0)),font=nf,fmt=mfmt)
            sc(ws2.cell(row=r,column=6,value=emp.get('employer_cost',0)),font=nf,fmt=mfmt)
            sc(ws2.cell(row=r,column=7,value=f'=E{r}+F{r}'),font=nf,fmt=mfmt)
            r+=1
        if emps:
            sc(ws2.cell(row=r,column=1,value=h.get('pay_date_hebrew','')),font=bf,fill=tfill)
            sc(ws2.cell(row=r,column=4,value='סה"כ תקופה'),font=bf,fill=tfill)
            for cc in [2,3]: sc(ws2.cell(row=r,column=cc),fill=tfill)
            sc(ws2.cell(row=r,column=5,value=f'=SUM(E{sr}:E{r-1})'),font=bf,fill=tfill,fmt=mfmt)
            sc(ws2.cell(row=r,column=6,value=f'=SUM(F{sr}:F{r-1})'),font=bf,fill=tfill,fmt=mfmt)
            sc(ws2.cell(row=r,column=7,value=f'=SUM(G{sr}:G{r-1})'),font=bf,fill=tfill,fmt=mfmt)
            inv_t=h.get('invoice_total',0)
            sc(ws2.cell(row=r,column=8,value=f'=IF(ABS(G{r}-{inv_t})<1,"✓","✗")'),font=nf,fill=tfill)
            r+=1
        r+=1
    for c,w in enumerate([14,12,12,22,18,18,18,10],1): ws2.column_dimensions[get_column_letter(c)].width=w

    # Sheet 3: סיכום עובד מצטבר
    ws3=wb.create_sheet('סיכום עובד מצטבר'); ws3.merge_cells('A1:G1')
    ws3['A1']=f'BRANDLIGHT INC. - Employee YTD Summary {yr}'; ws3['A1'].font=Font(name='Arial',bold=True,size=14,color='2F5496'); ws3['A1'].alignment=Alignment(horizontal='center')
    et={}
    for h in history:
        for emp in h.get('employees',[]):
            eid=emp.get('id','')
            if eid not in et: et[eid]={'name':emp['name'],'id':eid,'gross':0,'er':0,'periods':0}
            et[eid]['gross']+=emp.get('gross_pay',0); et[eid]['er']+=emp.get('employer_cost',0); et[eid]['periods']+=1
    for c,h in enumerate(['Employee ID','Employee Name','Periods','Total Gross ($)','Total ER Cost ($)','Total Cost ($)','% of Total'],1):
        sc(ws3.cell(row=3,column=c,value=h),font=hf,fill=hfill,align=Alignment(horizontal='center',wrap_text=True))
    gt=sum(e['gross']+e['er'] for e in et.values())
    r=4
    for eid in sorted(et,key=lambda x:et[x]['name']):
        e=et[eid]
        sc(ws3.cell(row=r,column=1,value=e['id']),font=nf)
        sc(ws3.cell(row=r,column=2,value=e['name']),font=nf)
        sc(ws3.cell(row=r,column=3,value=e['periods']),font=nf,align=Alignment(horizontal='center'))
        sc(ws3.cell(row=r,column=4,value=round(e['gross'],2)),font=nf,fmt=mfmt)
        sc(ws3.cell(row=r,column=5,value=round(e['er'],2)),font=nf,fmt=mfmt)
        sc(ws3.cell(row=r,column=6,value=f'=D{r}+E{r}'),font=bf,fmt=mfmt)
        sc(ws3.cell(row=r,column=7,value=(e['gross']+e['er'])/gt if gt else 0),font=nf,fmt=pfmt)
        r+=1
    sc(ws3.cell(row=r,column=1),fill=tfill); sc(ws3.cell(row=r,column=2,value='TOTAL'),font=bf,fill=tfill)
    sc(ws3.cell(row=r,column=3,value=f'=SUM(C4:C{r-1})'),font=bf,fill=tfill)
    for c in [4,5,6]: sc(ws3.cell(row=r,column=c,value=f'=SUM({get_column_letter(c)}4:{get_column_letter(c)}{r-1})'),font=bf,fill=tfill,fmt=mfmt)
    sc(ws3.cell(row=r,column=7,value=1),font=bf,fill=tfill,fmt=pfmt)
    for c,w in enumerate([14,22,10,20,22,20,12],1): ws3.column_dimensions[get_column_letter(c)].width=w

    # Sheet 4: התפלגות רכיבים
    ws4=wb.create_sheet('התפלגות רכיבים'); ws4.merge_cells('A1:D1')
    ws4['A1']=f'BRANDLIGHT INC. - Cost Component Distribution YTD {yr}'; ws4['A1'].font=Font(name='Arial',bold=True,size=14,color='2F5496'); ws4['A1'].alignment=Alignment(horizontal='center')
    for c,h in enumerate(['Component','Account #','YTD Total ($)','% of Total'],1):
        sc(ws4.cell(row=3,column=c,value=h),font=hf,fill=hfill,align=Alignment(horizontal='center',wrap_text=True))
    ct={}
    for h in history:
        comps=h.get('components',{})
        for field in INVOICE_FIELDS_ORDER: ct[field]=ct.get(field,0)+(comps.get(field,0) or 0)
    cg=sum(ct.values()); r=4
    for field in INVOICE_FIELDS_ORDER:
        an=FIELD_TO_ACCOUNT.get(field,field); num=DEFAULT_ACCOUNTS.get(an,""); val=ct.get(field,0)
        sc(ws4.cell(row=r,column=1,value=an),font=nf)
        sc(ws4.cell(row=r,column=2,value=num),font=nf,align=Alignment(horizontal='center'))
        sc(ws4.cell(row=r,column=3,value=val),font=nf,fmt=mfmt)
        sc(ws4.cell(row=r,column=4,value=val/cg if cg else 0),font=nf,fmt=pfmt)
        r+=1
    sc(ws4.cell(row=r,column=1,value='TOTAL'),font=bf,fill=tfill); sc(ws4.cell(row=r,column=2),fill=tfill)
    sc(ws4.cell(row=r,column=3,value=f'=SUM(C4:C{r-1})'),font=bf,fill=tfill,fmt=mfmt)
    sc(ws4.cell(row=r,column=4,value=1),font=bf,fill=tfill,fmt=pfmt)
    for c,w in enumerate([34,14,20,14],1): ws4.column_dimensions[get_column_letter(c)].width=w

    # Sheet 5: התפלגות חודשית
    ws5=wb.create_sheet('התפלגות חודשית'); ws5.merge_cells('A1:D1')
    ws5['A1']=f'BRANDLIGHT INC. - Monthly Distribution YTD {yr}'; ws5['A1'].font=Font(name='Arial',bold=True,size=14,color='2F5496'); ws5['A1'].alignment=Alignment(horizontal='center')
    for c,h in enumerate(['Month','Invoices','Total ($)','% of YTD'],1):
        sc(ws5.cell(row=3,column=c,value=h),font=hf,fill=hfill,align=Alignment(horizontal='center',wrap_text=True))
    mo={}
    for h in history:
        ds=h.get('pay_date','')
        if ds:
            p=ds.split('/');mk=f"{p[2]}-{p[0]}" if len(p)==3 else ds[:7];mn=datetime.strptime(f"{p[0]}/{p[2]}","%m/%Y").strftime("%B %Y") if len(p)==3 else ds[:7]
        else: mk=mn="Unknown"
        if mk not in mo: mo[mk]={'name':mn,'count':0,'total':0}
        mo[mk]['count']+=1; mo[mk]['total']+=h.get('invoice_total',0) or 0
    yt=sum(m['total'] for m in mo.values()); r=4
    for key in sorted(mo.keys()):
        m=mo[key]
        sc(ws5.cell(row=r,column=1,value=m['name']),font=nf)
        sc(ws5.cell(row=r,column=2,value=m['count']),font=nf,align=Alignment(horizontal='center'))
        sc(ws5.cell(row=r,column=3,value=m['total']),font=nf,fmt=mfmt)
        sc(ws5.cell(row=r,column=4,value=m['total']/yt if yt else 0),font=nf,fmt=pfmt)
        r+=1
    sc(ws5.cell(row=r,column=1,value='YTD TOTAL'),font=bf,fill=tfill)
    sc(ws5.cell(row=r,column=2,value=f'=SUM(B4:B{r-1})'),font=bf,fill=tfill,align=Alignment(horizontal='center'))
    sc(ws5.cell(row=r,column=3,value=f'=SUM(C4:C{r-1})'),font=bf,fill=tfill,fmt=mfmt)
    sc(ws5.cell(row=r,column=4,value=1),font=bf,fill=tfill,fmt=pfmt)
    for c,w in enumerate([20,12,20,14],1): ws5.column_dimensions[get_column_letter(c)].width=w
    return wb

# ============================================================
# UI
# ============================================================
st.title("📊 Brandlight Payroll → חשבשבת")
st.markdown("**העלי קבצי Payroll PDF → פקודת יומן + דוחות מצטברים**")
if 'history' not in st.session_state: st.session_state.history = load_history()

with st.sidebar:
    st.header("⚙️ חשבונות")
    accounts = {}
    for name,default in DEFAULT_ACCOUNTS.items(): accounts[name]=st.text_input(name,value=default,key=f"a_{name}")
    st.divider(); credit=st.text_input("חשבון זכות (ספק)",value=DEFAULT_CREDIT)
    st.divider(); st.header("📚 היסטוריה")
    st.metric("חשבוניות YTD",len(st.session_state.history))
    if st.session_state.history: st.metric("סה\"כ YTD",f"${sum(h.get('invoice_total',0) or 0 for h in st.session_state.history):,.2f}")
    if st.button("🗑️ נקה היסטוריה"): st.session_state.history=[]; save_history([]); st.rerun()

st.divider()
uploaded_files=st.file_uploader("📁 גררי קבצי US Payroll Consolidated PDF",type=['pdf'],accept_multiple_files=True)

if uploaded_files:
    st.info(f"📄 {len(uploaded_files)} קבצים נבחרו")
    all_data=[]; errors=[]; progress=st.progress(0)
    for i,f in enumerate(uploaded_files):
        progress.progress((i+1)/len(uploaded_files),text=f"מעבד: {f.name}")
        try:
            pd=extract_pdf_data(f)
            pd['pay_date_hebrew']=convert_date(pd['pay_date']) if pd['pay_date'] else ""
            if pd['invoice'] and pd['invoice'].get('total'): all_data.append(pd)
            else: errors.append(f"⚠️ {f.name}: לא נמצאו נתוני חשבונית")
        except Exception as e: errors.append(f"❌ {f.name}: {str(e)}")
    progress.empty()
    for err in errors: st.warning(err)
    if all_data:
        all_data.sort(key=lambda x:x.get('pay_date',''))
        st.subheader("📋 חשבוניות שזוהו")
        tbl=[]; bt=0
        for p in all_data:
            t=p['invoice'].get('total',0) or 0; bt+=t; comps=p['invoice'].get('components',{}); cs=sum(v for v in comps.values() if v)
            tbl.append({"תאריך":p['pay_date_hebrew'],"חשבונית":p['invoice_number'],"סה\"כ":f"${t:,.2f}","עובדים":len(p.get('employees',[])),"אימות":"✅" if abs(cs-t)<0.02 else "❌"})
        c1,c2=st.columns([3,1])
        with c1: st.dataframe(tbl,use_container_width=True,hide_index=True)
        with c2: st.metric("סה\"כ אצווה",f"${bt:,.2f}"); st.metric("חשבוניות",len(all_data)); st.metric("רשומות עובדים",sum(len(p.get('employees',[])) for p in all_data))

        if any(p.get('employees') for p in all_data):
            st.subheader("👥 עובדים שזוהו")
            ep=[]
            for p in all_data:
                for emp in p.get('employees',[]):
                    ep.append({"תאריך":p['pay_date_hebrew'],"מזהה":emp['id'],"שם":emp['name'],"ברוטו":f"${emp.get('gross_pay',0):,.2f}","עלות מעסיק":f"${emp.get('employer_cost',0):,.2f}","עלות כוללת":f"${emp.get('total_cost',0):,.2f}"})
            st.dataframe(ep,use_container_width=True,height=300,hide_index=True)

        td=sum((p['invoice'].get('components',{}).get(f,0) or 0) for p in all_data for f in INVOICE_FIELDS_ORDER)
        if abs(td-bt)<0.02: st.success(f"✅ מאוזן: ${td:,.2f}")
        else: st.error(f"❌ לא מאוזן! חובה ${td:,.2f} ≠ זכות ${bt:,.2f}")

        added=add_to_history(all_data,st.session_state.history)
        if added>0: save_history(st.session_state.history); st.toast(f"נוספו {added} חשבוניות")

        st.subheader("📥 הורדות")
        dc1,dc2=st.columns(2)
        with dc1:
            wb1=generate_journal(all_data,accounts,credit); o1=io.BytesIO(); wb1.save(o1); o1.seek(0)
            st.download_button("⬇️ פקודת יומן",data=o1,type="primary",file_name=f"Brandlight_Journal_{datetime.now().strftime('%Y%m%d')}.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        with dc2:
            if st.session_state.history:
                wb2=generate_summary(st.session_state.history); o2=io.BytesIO(); wb2.save(o2); o2.seek(0)
                st.download_button("📊 דוח מצטבר YTD",data=o2,file_name=f"Brandlight_YTD_{datetime.now().strftime('%Y%m%d')}.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
else:
    st.markdown("""
    ### 📌 הוראות שימוש
    1. **העלי קבצי PDF** — US Payroll Consolidated
    2. **בדקי נתונים** — אימות רכיבים + עובדים
    3. **הורידי Excel** — פקודת יומן + דוח מצטבר

    | רכיב | חשבון |
    |---|---|
    | Gross Wages | 800000 |
    | Expense Reimbursement | 800001 |
    | Employer Fed & State Taxes | 800002 |
    | Workers Compensation | 800003 |
    | Employee Benefits | 800004 |
    | Administrative Fee | 800005 |
    | 401(k) ER Contribution | 800006 |
    | 401(k) Establishment Fee | 800007 |
    | **חשבון זכות** | **540001** |
    """)

if st.session_state.history:
    st.divider(); st.subheader("📚 דוחות מצטברים YTD")
    tc1,tc2=st.columns([2,1])
    with tc1:
        hd=[{"תאריך":h.get('pay_date_hebrew',''),"חשבונית":h.get('invoice_number',''),"עובדים":len(h.get('employees',[])),"סה\"כ":f"${h.get('invoice_total',0):,.2f}"} for h in st.session_state.history]
        st.dataframe(hd,use_container_width=True,hide_index=True)
    with tc2:
        st.metric("חשבוניות",len(st.session_state.history))
        st.metric("סה\"כ YTD",f"${sum(h.get('invoice_total',0) or 0 for h in st.session_state.history):,.2f}")
        wb_s=generate_summary(st.session_state.history); os_b=io.BytesIO(); wb_s.save(os_b); os_b.seek(0)
        st.download_button("📊 הורד דוח מצטבר YTD",data=os_b,type="primary",file_name=f"Brandlight_YTD_Summary_{datetime.now().strftime('%Y%m%d')}.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
